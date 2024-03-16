import os
import threading
import uuid
from typing import Union

from openpyxl.reader.excel import load_workbook
from typing import Generator

from openpyxl.workbook import Workbook
from .Exceptions import *

from constants import SHEETS_FILES_LOCATION


class SheetLoader:
    _instance = None

    def __new__(cls, *args, **kwargs):
        if not cls._instance:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self, directory=SHEETS_FILES_LOCATION):
        if not hasattr(self, 'sheets'):
            self.sheets = {}
        if not hasattr(self, 'locks'):
            self.locks = {}
        if not hasattr(self, 'directory'):
            self.directory = directory

    def create(self, columns: list):
        try:
            id = str(uuid.uuid4())
            path = os.path.join(self.directory, f"{id}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(columns)
            wb.save(path)
            self.sheets[id] = path
            self.locks[id] = threading.Lock()
            return id
        except Exception as e:
            raise SheetCreationException

    def get_sheet_path_by_id(self, sheet_id: str):
        if not self.sheets:
            self.load_sheets()
        if self.locks.get(sheet_id) is None:
            return None
        with self.locks[sheet_id]:
            return self.sheets.get(sheet_id)

    def load_sheets(self):
        files = os.listdir(self.directory)
        for file in files:
            self.sheets[file.split('.')[0]] = os.path.join(self.directory, file)
            self.locks[file.split('.')[0]] = threading.Lock()

    def update_sheet_cell(self, sheet_id: str, row_index: int, column_index: int, value: str):
        try:
            path = self.get_sheet_path_by_id(sheet_id)
            with self.locks[sheet_id]:
                wb = load_workbook(path)
                ws = wb.active
                ws.cell(row=row_index + 1, column=column_index + 1).value = value
                wb.save(path)
        except Exception as e:
            raise UpdateSheetCellException()

    def get_cell_value(self, sheet_id: str, column_name: str, row_index: int) -> Union[int, None]:
        sheet_path = self.get_sheet_path_by_id(sheet_id)
        if sheet_path is None:
            raise SheetGetValueException()

        try:
            with self.locks[sheet_id]:
                wb = load_workbook(sheet_path)
                ws = wb.active
                column_names = [col.value for col in ws[1]]
                column_index = column_names.index(column_name)
                return ws.cell(row=row_index + 1, column=column_index + 1).value
        except Exception:
            raise SheetGetValueException()

    def get_sheet_generator(self, sheet_id: str) -> Generator:
        wb = load_workbook(self.get_sheet_path_by_id(sheet_id))
        ws = wb.active
        return ws.iter_rows(values_only=True)
